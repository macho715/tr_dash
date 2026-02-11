"""
Build ADNOC_DURATION_comparison.xlsx with sheets: Comparison, MAPPING LOG.
Requires: ADNOC_DURATION_comparison.csv and ADNOC_DURATION_mapping_log.json (from build_adnoc_comparison.js).
Uses openpyxl (pip install openpyxl).
"""
import csv
import json
import os
from pathlib import Path

try:
    from openpyxl import Workbook
except ImportError:
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook

BASE = Path(__file__).resolve().parent
CSV_PATH = BASE / "ADNOC_DURATION_comparison.csv"
JSON_PATH = BASE / "ADNOC_DURATION_mapping_log.json"
XLSX_PATH = BASE / "ADNOC_DURATION_comparison.xlsx"


def read_csv_rows(path):
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        return list(reader)


def main():
    if not CSV_PATH.exists():
        print("Run build_adnoc_comparison.js first to create", CSV_PATH.name)
        return 1
    if not JSON_PATH.exists():
        print("Run build_adnoc_comparison.js first to create", JSON_PATH.name)
        return 1

    rows = read_csv_rows(CSV_PATH)
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        mapping = json.load(f)

    wb = Workbook()

    # Sheet 1: Comparison (from CSV)
    ws_comp = wb.active
    ws_comp.title = "Comparison"
    for i, row in enumerate(rows, 1):
        for j, val in enumerate(row, 1):
            ws_comp.cell(row=i, column=j, value=val)

    # Sheet 2: MAPPING LOG
    ws_log = wb.create_sheet("MAPPING LOG", 1)

    # Block 1: WA in Phase
    ws_log.cell(row=1, column=1, value="WA in Phase")
    wa_headers = ["phase_id", "event_id", "occurred_date", "occurred_time", "wa_dt"]
    for j, h in enumerate(wa_headers, 1):
        ws_log.cell(row=2, column=j, value=h)
    for i, r in enumerate(mapping["wa"], 3):
        ws_log.cell(row=i, column=1, value=r.get("phase_id"))
        ws_log.cell(row=i, column=2, value=r.get("event_id"))
        ws_log.cell(row=i, column=3, value=r.get("occurred_date"))
        ws_log.cell(row=i, column=4, value=r.get("occurred_time"))
        ws_log.cell(row=i, column=5, value=r.get("wa_dt"))

    wa_end = 2 + len(mapping["wa"])
    next_row = wa_end + 2

    # Block 2: Bushra overlap
    ws_log.cell(row=next_row, column=1, value="Bushra overlap by phase")
    next_row += 1
    bushra_headers = ["phase_id", "bushra_no", "bushra_start_dt", "bushra_end_dt", "overlap_hrs"]
    for j, h in enumerate(bushra_headers, 1):
        ws_log.cell(row=next_row, column=j, value=h)
    next_row += 1
    for r in mapping["bushra"]:
        ws_log.cell(row=next_row, column=1, value=r.get("phase_id"))
        ws_log.cell(row=next_row, column=2, value=r.get("bushra_no"))
        ws_log.cell(row=next_row, column=3, value=r.get("bushra_start_dt"))
        ws_log.cell(row=next_row, column=4, value=r.get("bushra_end_dt"))
        ws_log.cell(row=next_row, column=5, value=r.get("overlap_hrs"))
        next_row += 1

    next_row += 1

    # Block 3: Activity (phase → activity_id)
    ws_log.cell(row=next_row, column=1, value="Activity (phase → activity_id)")
    next_row += 1
    act_headers = ["phase_id", "activity_id_primary", "activity_id_secondary"]
    for j, h in enumerate(act_headers, 1):
        ws_log.cell(row=next_row, column=j, value=h)
    next_row += 1
    for r in mapping["activity"]:
        ws_log.cell(row=next_row, column=1, value=r.get("phase_id"))
        ws_log.cell(row=next_row, column=2, value=r.get("activity_id_primary"))
        ws_log.cell(row=next_row, column=3, value=r.get("activity_id_secondary"))
        next_row += 1

    wb.save(XLSX_PATH)
    print("Created", XLSX_PATH)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
