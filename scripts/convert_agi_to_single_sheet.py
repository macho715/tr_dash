#!/usr/bin/env python3
"""
Convert AGI TR Final Schedule JSON to single-sheet Excel with TR grouping.
"""

import json
import sys
from pathlib import Path
from datetime import datetime

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("ERROR: Required packages not found.")
    print("Install with: pip install pandas openpyxl")
    sys.exit(1)


# TR Unit 색상 매핑
TR_COLORS = {
    'Mobilization': 'F5F5F5',
    'TR Unit 1': 'E3F2FD',
    'TR Unit 2': 'E8F5E9',
    'TR Unit 3': 'FFF9C4',
    'TR Unit 4': 'FFE0B2',
    'TR Unit 5': 'F3E5F5',
    'TR Unit 6': 'FCE4EC',
    'TR Unit 7': 'E0F2F1',
    'Demobilization': 'F5F5F5',
}

# 헤더 색상 (진한 버전)
TR_HEADER_COLORS = {
    'Mobilization': 'BDBDBD',
    'TR Unit 1': '90CAF9',
    'TR Unit 2': 'A5D6A7',
    'TR Unit 3': 'FFF59D',
    'TR Unit 4': 'FFCC80',
    'TR Unit 5': 'CE93D8',
    'TR Unit 6': 'F48FB1',
    'TR Unit 7': 'B2DFDB',
    'Demobilization': 'BDBDBD',
}


def assign_tr_unit(activity):
    """활동에 TR Unit 할당"""
    level1 = activity.get('level1', '')
    level2 = activity.get('level2', '')
    activity_id = activity.get('activity_id')
    
    if level1 == 'MOBILIZATION':
        return 'Mobilization'
    elif level1 == 'DEMOBILIZATION':
        return 'Demobilization'
    elif level2 and 'AGI TR Unit' in level2:
        # "AGI TR Unit 1" -> "TR Unit 1"
        return level2.replace('AGI ', '')
    elif activity_id:
        # ID로 판단
        if activity_id.startswith('A1'):
            return 'TR Unit 1'
        elif activity_id.startswith('A2'):
            # A2000대는 TR Unit 2-7
            id_num = int(activity_id[1:])
            if 2000 <= id_num < 2100:
                return 'Demobilization' if id_num in [2000, 2010, 2020] else 'TR Unit 2'
            elif 2100 <= id_num < 2200:
                return 'TR Unit 2'
            elif 2200 <= id_num < 2400:
                return 'TR Unit 3'
            elif 2400 <= id_num < 2600:
                return 'TR Unit 4'
            elif 2600 <= id_num < 2700:
                return 'TR Unit 5'
            elif 2700 <= id_num < 2800:
                return 'TR Unit 6'
            else:
                return 'TR Unit 7'
    
    return 'Other'


def load_and_process_json(json_path):
    """JSON 로드 및 처리"""
    print(f"Loading {json_path}...")
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    activities = data.get('activities', [])
    print(f"Found {len(activities)} activities")
    
    # TR Unit 할당 및 데이터 준비
    rows = []
    for activity in activities:
        activity_id = activity.get('activity_id')
        
        # Summary 행 (activity_id가 null)은 제외
        if activity_id is None:
            continue
        
        tr_unit = assign_tr_unit(activity)
        
        row = {
            'TR Unit': tr_unit,
            'Date': activity.get('planned_start', ''),
            'Activity ID': activity_id,
            'Activity Name': activity.get('activity_name', ''),
            'Duration': activity.get('duration', ''),
            'Start Date': activity.get('planned_start', ''),
            'End Date': activity.get('planned_finish', ''),
            'Level 1': activity.get('level1', ''),
            'Level 2': activity.get('level2', ''),
        }
        rows.append(row)
    
    # DataFrame 생성
    df = pd.DataFrame(rows)
    
    # TR Unit 순서 정의
    tr_order = [
        'Mobilization',
        'TR Unit 1',
        'TR Unit 2',
        'TR Unit 3',
        'TR Unit 4',
        'TR Unit 5',
        'TR Unit 6',
        'TR Unit 7',
        'Demobilization'
    ]
    
    # TR Unit을 카테고리로 변환하여 정렬
    df['TR Unit'] = pd.Categorical(df['TR Unit'], categories=tr_order, ordered=True)
    
    # 정렬: TR Unit 순서 -> 날짜 순서
    df = df.sort_values(['TR Unit', 'Date', 'Activity ID'])
    
    print(f"Processed {len(df)} activities")
    
    return df, data


def create_excel_with_styling(df, output_path, metadata):
    """스타일이 적용된 엑셀 생성"""
    print("Creating Excel workbook...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "All Activities"
    
    # 헤더 작성
    headers = list(df.columns)
    ws.append(headers)
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 현재 TR Unit 추적
    current_tr = None
    current_row = 2
    
    for idx, row in df.iterrows():
        tr_unit = row['TR Unit']
        
        # TR Unit이 변경되면 구분 헤더 행 추가
        if tr_unit != current_tr:
            current_tr = tr_unit
            
            # 구분 헤더 행
            header_row = [tr_unit] + [''] * (len(headers) - 1)
            ws.append(header_row)
            
            # 구분 헤더 스타일
            header_color = TR_HEADER_COLORS.get(tr_unit, 'CCCCCC')
            header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
            header_font = Font(bold=True, size=12)
            thick_border = Border(
                top=Side(style='thick'),
                bottom=Side(style='thick')
            )
            
            for cell in ws[current_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = thick_border
            
            # 셀 병합 (TR Unit 이름)
            ws.merge_cells(f'A{current_row}:{chr(65+len(headers)-1)}{current_row}')
            
            current_row += 1
        
        # 데이터 행 추가
        data_row = [
            row['TR Unit'],
            row['Date'],
            row['Activity ID'],
            row['Activity Name'],
            row['Duration'],
            row['Start Date'],
            row['End Date'],
            row['Level 1'],
            row['Level 2']
        ]
        ws.append(data_row)
        
        # 데이터 행 스타일
        bg_color = TR_COLORS.get(tr_unit, 'FFFFFF')
        bg_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        
        for cell in ws[current_row]:
            cell.fill = bg_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        current_row += 1
    
    # 열 너비 조정
    column_widths = {
        'A': 15,  # TR Unit
        'B': 12,  # Date
        'C': 12,  # Activity ID
        'D': 60,  # Activity Name
        'E': 10,  # Duration
        'F': 12,  # Start Date
        'G': 12,  # End Date
        'H': 15,  # Level 1
        'I': 20,  # Level 2
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # 동결 창 (헤더)
    ws.freeze_panes = 'A2'
    
    # 필터 활성화
    ws.auto_filter.ref = ws.dimensions
    
    # 저장
    print(f"Saving to {output_path}...")
    wb.save(output_path)
    
    file_size = output_path.stat().st_size / 1024
    print(f"OK Excel file created successfully!")
    print(f"  File size: {file_size:.1f} KB")
    print(f"  Total rows: {current_row - 1} (including headers)")


def main():
    """메인 실행"""
    json_path = Path("data/schedule/agi tr final schedule.json")
    output_path = Path("data/schedule/agi_tr_final_schedule.xlsx")
    
    if not json_path.exists():
        print(f"ERROR: Input file not found: {json_path}")
        sys.exit(1)
    
    try:
        # JSON 처리
        df, metadata = load_and_process_json(json_path)
        
        # 엑셀 생성
        create_excel_with_styling(df, output_path, metadata)
        
        # 통계 출력
        print("\n[SUCCESS] Conversion complete!")
        print(f"   Output: {output_path}")
        print(f"\nTR Unit Statistics:")
        tr_counts = df['TR Unit'].value_counts().sort_index()
        for tr_unit, count in tr_counts.items():
            print(f"   {tr_unit}: {count} activities")
        
    except Exception as e:
        print(f"\n[ERROR] Conversion failed: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
