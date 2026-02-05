#!/usr/bin/env python3
"""
Convert option_c_v0.8.0.json to Excel format with multiple sheets.

Usage:
    python scripts/json_to_excel.py
"""

import json
import sys
from pathlib import Path
from datetime import datetime

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("ERROR: Required packages not found.")
    print("Install with: pip install pandas openpyxl")
    sys.exit(1)


def load_json(json_path: Path) -> dict:
    """Load JSON file."""
    print(f"Loading {json_path}...")
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def flatten_activities(activities_dict: dict) -> pd.DataFrame:
    """Flatten activities dictionary to DataFrame."""
    rows = []
    
    for activity_id, activity in activities_dict.items():
        row = {
            'activity_id': activity_id,
            'activity_name': activity.get('activity_name', ''),
            'level1': activity.get('level1', ''),
            'level2': activity.get('level2', ''),
            'status': activity.get('status', ''),
            'state': activity.get('state', ''),
            'planned_start': activity.get('planned_start', ''),
            'planned_finish': activity.get('planned_finish', ''),
            'actual_start': activity.get('actual_start', ''),
            'actual_finish': activity.get('actual_finish', ''),
            'duration': activity.get('duration', ''),
            'priority': activity.get('priority', ''),
            'lock_level': activity.get('lock_level', ''),
            'is_locked': activity.get('is_locked', False),
            'blocker_code': activity.get('blocker_code', ''),
            'anchor_type': activity.get('anchor_type', ''),
            'constraint': activity.get('constraint', ''),
            'resource_tags': ', '.join(activity.get('resource_tags', [])),
            'evidence_ids': ', '.join(activity.get('evidence_ids', [])),
            'trip_id': activity.get('trip_id', ''),
        }
        
        # Dependencies
        depends_on = activity.get('depends_on', [])
        if depends_on:
            dep_list = []
            for dep in depends_on:
                pred_id = dep.get('predecessorId', '')
                dep_type = dep.get('type', 'FS')
                lag = dep.get('lagDays', 0)
                dep_list.append(f"{pred_id}({dep_type}{'+' if lag > 0 else ''}{lag if lag != 0 else ''})")
            row['depends_on'] = ', '.join(dep_list)
        else:
            row['depends_on'] = ''
        
        rows.append(row)
    
    return pd.DataFrame(rows)


def flatten_trips(trips_dict: dict) -> pd.DataFrame:
    """Flatten trips dictionary to DataFrame."""
    rows = []
    
    for trip_id, trip in trips_dict.items():
        rows.append({
            'trip_id': trip_id,
            'name': trip.get('name', ''),
            'sequence': trip.get('sequence', ''),
            'tr_ids': ', '.join(trip.get('tr_ids', [])),
            'activity_count': len(trip.get('activity_ids', [])),
            'activity_ids': ', '.join(trip.get('activity_ids', []))
        })
    
    return pd.DataFrame(rows)


def flatten_resources(resources_dict: dict) -> pd.DataFrame:
    """Flatten resources dictionary to DataFrame."""
    rows = []
    
    for resource_id, resource in resources_dict.items():
        rows.append({
            'resource_id': resource_id,
            'name': resource.get('name', ''),
            'type': resource.get('type', ''),
            'capacity': resource.get('capacity', ''),
            'available_from': resource.get('available_from', ''),
            'available_to': resource.get('available_to', ''),
            'tags': ', '.join(resource.get('tags', []))
        })
    
    return pd.DataFrame(rows)


def style_header_row(ws, row_num: int):
    """Apply header styling."""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        bottom=Side(style='thick', color='366092')
    )
    
    for cell in ws[row_num]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border


def style_sheet(ws, df: pd.DataFrame, freeze_panes: str = 'A2'):
    """Apply general styling to worksheet."""
    # Freeze panes
    ws.freeze_panes = freeze_panes
    
    # Style header
    style_header_row(ws, 1)
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Set width with min/max bounds
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Alternate row colors
    alt_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    for row_num in range(2, len(df) + 2):
        if row_num % 2 == 0:
            for cell in ws[row_num]:
                if not cell.fill.start_color.rgb or cell.fill.start_color.rgb == '00000000':
                    cell.fill = alt_fill


def create_summary_sheet(wb: Workbook, data: dict):
    """Create summary overview sheet."""
    ws = wb.create_sheet("Summary", 0)
    
    # Title
    ws['A1'] = "TR Dashboard SSOT - Summary"
    ws['A1'].font = Font(bold=True, size=16, color="366092")
    ws.merge_cells('A1:D1')
    
    # Contract info
    contract = data.get('contract', {})
    ws['A3'] = "Contract Information"
    ws['A3'].font = Font(bold=True, size=12)
    ws['A4'] = "Name:"
    ws['B4'] = contract.get('name', '')
    ws['A5'] = "Version:"
    ws['B5'] = contract.get('version', '')
    ws['A6'] = "Timezone:"
    ws['B6'] = contract.get('timezone', '')
    
    # Entity counts
    entities = data.get('entities', {})
    ws['A8'] = "Entity Counts"
    ws['A8'].font = Font(bold=True, size=12)
    ws['A9'] = "Activities:"
    ws['B9'] = len(entities.get('activities', {}))
    ws['A10'] = "Trips:"
    ws['B10'] = len(entities.get('trips', {}))
    ws['A11'] = "Resources:"
    ws['B11'] = len(entities.get('resources', {}))
    
    # Enums summary
    catalog = data.get('catalog', {})
    enums = catalog.get('enums', {})
    ws['A13'] = "Activity States"
    ws['A13'].font = Font(bold=True, size=12)
    states = enums.get('activity_state', [])
    for idx, state in enumerate(states, start=14):
        ws[f'A{idx}'] = state
    
    ws['C13'] = "Lock Levels"
    ws['C13'].font = Font(bold=True, size=12)
    locks = enums.get('lock_level', [])
    for idx, lock in enumerate(locks, start=14):
        ws[f'C{idx}'] = lock
    
    # Adjust columns
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    
    # Generated timestamp
    ws['A25'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A25'].font = Font(italic=True, size=9, color="666666")


def convert_json_to_excel(json_path: Path, output_path: Path):
    """Main conversion function."""
    # Load JSON
    data = load_json(json_path)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create summary sheet
    print("Creating summary sheet...")
    create_summary_sheet(wb, data)
    
    # Extract entities
    entities = data.get('entities', {})
    
    # Activities sheet
    if 'activities' in entities and entities['activities']:
        print("Converting activities...")
        df_activities = flatten_activities(entities['activities'])
        ws_activities = wb.create_sheet("Activities")
        
        for r in dataframe_to_rows(df_activities, index=False, header=True):
            ws_activities.append(r)
        
        style_sheet(ws_activities, df_activities)
        print(f"  OK {len(df_activities)} activities")
    
    # Trips sheet
    if 'trips' in entities and entities['trips']:
        print("Converting trips...")
        df_trips = flatten_trips(entities['trips'])
        ws_trips = wb.create_sheet("Trips")
        
        for r in dataframe_to_rows(df_trips, index=False, header=True):
            ws_trips.append(r)
        
        style_sheet(ws_trips, df_trips)
        print(f"  OK {len(df_trips)} trips")
    
    # Resources sheet
    if 'resources' in entities and entities['resources']:
        print("Converting resources...")
        df_resources = flatten_resources(entities['resources'])
        ws_resources = wb.create_sheet("Resources")
        
        for r in dataframe_to_rows(df_resources, index=False, header=True):
            ws_resources.append(r)
        
        style_sheet(ws_resources, df_resources)
        print(f"  OK {len(df_resources)} resources")
    
    # Save workbook
    print(f"\nSaving to {output_path}...")
    wb.save(output_path)
    print(f"OK Excel file created successfully!")
    print(f"  File size: {output_path.stat().st_size / 1024:.1f} KB")


def main():
    """Main entry point."""
    # Define paths
    project_root = Path(__file__).parent.parent
    json_path = project_root / "data" / "schedule" / "option_c_v0.8.0.json"
    output_path = project_root / "data" / "schedule" / "option_c_v0.8.0.xlsx"
    
    # Verify input exists
    if not json_path.exists():
        print(f"ERROR: Input file not found: {json_path}")
        sys.exit(1)
    
    # Convert
    try:
        convert_json_to_excel(json_path, output_path)
        print(f"\n[SUCCESS] Conversion complete!")
        print(f"   Output: {output_path}")
    except Exception as e:
        print(f"\n[ERROR] Conversion failed: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
